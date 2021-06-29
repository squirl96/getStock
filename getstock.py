from pathlib import Path
import csv
import requests, re, sys #, time
from bs4 import BeautifulSoup
import gettext
from time import sleep
# import xlrd #for reading xlsx file
import pandas as pd
# from tqdm import tqdm
import openpyxl # load_workbook


def GetData(url):
    # Get the datas from the Digikey page
    head = requests.utils.default_headers()
    head.update({ 'User-Agent': 'Mozilla/5.0 (X11; Ubuntu: Linux x86_64; rv:52.0) Gecko/20100101 Firefox/52.0'})
    req = requests.get(url, head, timeout=5)
    soup = BeautifulSoup(req.content, 'html.parser')
    return(soup)


def main():
    # load excel with its path
    wrkbk = openpyxl.load_workbook("cl.xlsx")

    sh = wrkbk.active
    sp = 20 - len("Name")
    print("Name" + sp*" " + "DG stock " + "Mouser Stock")

    # iterate through excel and display data
    for row in sh.iter_rows(min_row=3, values_only=True):
        name = row[0]
        urlDG = row[1]
        urlM = row[2]

        # ----------------   Get Digikey's stock   ----------------
        Data = GetData(urlDG)
        table = Data.find(lambda tag: tag.name =='div' and tag.has_attr('id') and tag['id']=="quantityAvailable")
        avail = table.find(lambda tag: tag.name == 'span' and tag.has_attr('id') and tag['id']=="dkQty")

        fin = str(avail)

        DGstock = ""
        text = fin.find("dkQty")
        c = text + len("dkQty") + 2
        while fin[c] != "<":
            DGstock = DGstock + fin[c]
            c = c + 1


        # ----------------   Get Mouser's stock   -----------------
        # Data = GetData(urlM)
        # table = Data.find(lambda tag: tag.name =='div' and tag.has_attr('class') and tag['class']=="col-xs-3 onOrderQuantity")
        # # avail = table.find(lambda tag: tag.name == 'span' and tag.has_attr('id') and tag['id']=="dkQty")
        # print(table)



        sp = 20 - len(name)
        sp2 = 9 - len(DGstock)
        print(name + sp*" " + DGstock + sp2*" " + "a")
        sleep(10)
    # print(df)
    # stock = re.sub(r'(\r\n *)', "", Data.find("Non-Stock").gettext())

    # if description != 0:
    #     print("not ok")
    # else:
    #     print("help")

#------------------------------------------------------------------------------

if __name__=="__main__":
    main()
